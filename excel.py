from http.server import BaseHTTPRequestHandler
import json, io, datetime, base64
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

CORS = {
    'Access-Control-Allow-Origin': '*',
    'Access-Control-Allow-Methods': 'POST, OPTIONS',
    'Access-Control-Allow-Headers': 'Content-Type',
}

class handler(BaseHTTPRequestHandler):

    def do_OPTIONS(self):
        self.send_response(200)
        for k,v in CORS.items(): self.send_header(k, v)
        self.end_headers()

    def do_POST(self):
        try:
            length = int(self.headers.get('Content-Length', 0))
            body   = json.loads(self.rfile.read(length))
            wb     = generar_excel(
                body.get('vendedor',''), body.get('mes',''),
                int(body.get('año',2026)),
                body.get('detalle',[]), body.get('resumen',{})
            )
            buf  = io.BytesIO()
            wb.save(buf)
            b64  = base64.b64encode(buf.getvalue()).decode('ascii')
            nombre = f"Liquidacion-{body.get('vendedor','').split(',')[0].strip()}-{body.get('mes','')}-{body.get('año',2026)}.xlsx"
            resp = json.dumps({'data': b64, 'filename': nombre})

            self.send_response(200)
            self.send_header('Content-Type', 'application/json')
            self.send_header('Content-Length', str(len(resp)))
            for k,v in CORS.items(): self.send_header(k, v)
            self.end_headers()
            self.wfile.write(resp.encode())

        except Exception as e:
            self.send_response(500)
            self.send_header('Content-Type', 'application/json')
            for k,v in CORS.items(): self.send_header(k, v)
            self.end_headers()
            self.wfile.write(json.dumps({'error': str(e)}).encode())


def st(bold=False, size=11, color=None):
    kw = {'name':'Arial','bold':bold,'size':size}
    if color: kw['color'] = color
    return Font(**kw)

def fl(c): return PatternFill(patternType='solid', fgColor=c)
def al(h='center',v='center',wrap=False): return Alignment(horizontal=h,vertical=v,wrap_text=wrap)
def bd():
    t = Side(style='thin', color='000000')
    return Border(left=t,right=t,top=t,bottom=t)

def cel(ws, coord, val=None, font=None, fill=None, align=None, border=None, fmt=None):
    c = ws[coord]
    if val is not None: c.value = val
    if font:   c.font      = font
    if fill:   c.fill      = fill
    if align:  c.alignment = align
    if border: c.border    = border
    if fmt:    c.number_format = fmt
    return c


def generar_excel(vendedor, mes, año, detalle, resumen):
    wb = Workbook()
    ws = wb.active
    ws.title = 'ESQUEMA COMISIONAL'

    FT=st(True,12); FH=st(True,10); FD=st(False,9)
    FR=st(True,9,'FF0000'); FG=st(True,9,'00B050'); FB=st(True,9)
    FTOPE=fl('FF9999'); FHEAD=fl('D9D9D9')
    AC=al(); BRD=bd()

    # Formats
    FMT_PESOS = '"$"#,##0'           # $59.383.000
    FMT_PCT   = '0.00%'              # 0.90%
    FMT_NUM   = '#,##0'
    FMT_DATE  = 'DD/MM/YYYY'

    cel(ws,'B2','ESQUEMA COMISIONAL',font=FT,align=al('left'))
    cel(ws,'B3','COLABORADOR',font=FT)
    cel(ws,'D3',vendedor.upper().replace(', ',' '),font=FT,align=AC)
    cel(ws,'B4','MES DE COMISION',font=FT)

    MESES=['ENERO','FEBRERO','MARZO','ABRIL','MAYO','JUNIO',
           'JULIO','AGOSTO','SEPTIEMBRE','OCTUBRE','NOVIEMBRE','DICIEMBRE']
    mn = MESES.index(mes.upper())+1 if mes.upper() in MESES else 1
    c=ws['D4']; c.value=datetime.date(año,mn,1)
    c.font=FT; c.alignment=AC; c.number_format='MMMM YYYY'

    cel(ws,'Q5','TOPE 1,5%',font=st(True,11),fill=FTOPE,align=AC)

    # Headers — now with Compartida in col C and Solicitud in E
    hdrs=[('B','Cantidad'),('C','Compartida'),('D','Fecha'),('E','Solicitud'),
          ('F','Titular'),('G','Modelo'),('H','Valor Vehiculo'),('I','ESCALA'),
          ('J','DESCUENTO CTA 1'),('K','ESCALA A (70%) -0,10%'),
          ('L','ESCALA B (90-100%) -0,15%'),('M','Lead (+0,15%)'),
          ('N','Objetivo (+0,15%)'),('O','Compartida (div)'),('P','Sub Total'),
          ('Q','Comision Final'),('R','Remuneracion')]
    for col,val in hdrs:
        cel(ws,f'{col}6',val,font=FH,fill=FHEAD,align=al('center',wrap=True),border=BRD)

    ws.column_dimensions['B'].width=8
    ws.column_dimensions['C'].width=18  # Compartida name
    ws.column_dimensions['D'].width=13
    ws.column_dimensions['E'].width=12
    ws.column_dimensions['F'].width=28
    ws.column_dimensions['G'].width=24
    ws.column_dimensions['H'].width=16
    ws.row_dimensions[6].height=30
    for col in ['I','J','K','L','M','N','O','P','Q','R']:
        ws.column_dimensions[col].width=12

    NUM=max(18,len(detalle)+2)
    pct=float(resumen.get('pctBase',0.0125))
    has_lead=bool(resumen.get('hasLead',False))
    cumple=bool(resumen.get('cumpleObj',False))

    for i in range(NUM):
        r=7+i
        # B: row number
        cel(ws,f'B{r}',i+1,font=FD,align=AC,border=BRD)
        # O: compartida divisor (1 = full, used in formula)
        cel(ws,f'O{r}',1,font=FB,align=AC,border=BRD)
        # Formulas with percentage format
        cel(ws,f'P{r}',f'=+(I{r}-K{r}-L{r}+M{r}+N{r})/O{r}',
            font=FB,align=AC,border=BRD,fmt=FMT_PCT)
        cel(ws,f'Q{r}',f'=MIN(P{r},0.015)',
            font=FB,align=AC,border=BRD,fmt=FMT_PCT)
        cel(ws,f'R{r}',f'=+H{r}*Q{r}',
            font=FD,align=AC,border=BRD,fmt=FMT_PESOS)

        if i<len(detalle):
            d=detalle[i]
            bonif=float(d.get('bonif',0) or 0)
            prot=bool(d.get('protegida',False))
            compartida=str(d.get('compartida','') or '').strip()
            es_compartida = compartida and compartida not in ['-','nan','']

            # C: nombre del compartidor
            cel(ws,f'C{r}',compartida if es_compartida else '',
                font=FD,align=AC,border=BRD)

            # If shared, divisor = 2
            if es_compartida:
                ws[f'O{r}'].value = 2

            # D: fecha
            fecha=d.get('fecha','')
            if fecha:
                try:
                    dt=datetime.datetime.strptime(fecha[:10],'%Y-%m-%d')
                    c2=ws[f'D{r}']; c2.value=dt; c2.font=FD
                    c2.alignment=AC; c2.border=BRD; c2.number_format=FMT_DATE
                except: cel(ws,f'D{r}',fecha,font=FD,align=AC,border=BRD)
            else: cel(ws,f'D{r}','',font=FD,align=AC,border=BRD)

            # E: numero de solicitud REAL
            sol=str(d.get('solicitud','') or '').replace('.0','').strip()
            sol_val = int(float(sol)) if sol and sol not in ['nan',''] else ''
            cel(ws,f'E{r}',sol_val,font=FD,align=AC,border=BRD)

            cel(ws,f'F{r}',(d.get('cliente','') or '').upper(),font=FD,align=AC,border=BRD)
            cel(ws,f'G{r}',d.get('modelo','') or '',font=FD,align=AC,border=BRD)

            # H: valor en pesos
            cel(ws,f'H{r}',float(d.get('vl',0) or 0),
                font=FD,align=AC,border=BRD,fmt=FMT_PESOS)

            # I: escala como porcentaje
            cel(ws,f'I{r}',pct,font=FD,align=AC,border=BRD,fmt=FMT_PCT)

            # J: bonificacion
            cel(ws,f'J{r}',bonif,font=FD,align=AC,border=BRD,fmt='0%')

            # K,L: descuentos como porcentaje
            escA=0 if prot else (0.001 if 0.699<=bonif<0.9 else 0)
            escB=0 if prot else (0.0015 if bonif>=0.9 else 0)
            cel(ws,f'K{r}',escA,font=FR,align=AC,border=BRD,fmt=FMT_PCT)
            cel(ws,f'L{r}',escB,font=FR,align=AC,border=BRD,fmt=FMT_PCT)
            cel(ws,f'M{r}',0.0015 if has_lead else 0,font=FG,align=AC,border=BRD,fmt=FMT_PCT)
            cel(ws,f'N{r}',0.0015 if cumple else 0,font=FG,align=AC,border=BRD,fmt=FMT_PCT)
        else:
            # Empty rows
            cel(ws,f'C{r}','',font=FD,align=AC,border=BRD)
            for col in ['D','E','F','G','H']:
                cel(ws,f'{col}{r}',None,font=FD,align=AC,border=BRD)
            cel(ws,f'I{r}',pct,font=FD,align=AC,border=BRD,fmt=FMT_PCT)
            cel(ws,f'J{r}',0,font=FD,align=AC,border=BRD,fmt='0%')
            cel(ws,f'K{r}',0,font=FR,align=AC,border=BRD,fmt=FMT_PCT)
            cel(ws,f'L{r}',0,font=FR,align=AC,border=BRD,fmt=FMT_PCT)
            cel(ws,f'M{r}',0.0015 if has_lead else 0,font=FG,align=AC,border=BRD,fmt=FMT_PCT)
            cel(ws,f'N{r}',0.0015 if cumple else 0,font=FG,align=AC,border=BRD,fmt=FMT_PCT)

    # Total row
    tr=7+NUM
    cel(ws,f'R{tr}',f'=SUM(R7:R{tr-1})',font=st(True,11),border=BRD,fmt=FMT_PESOS)

    return wb
